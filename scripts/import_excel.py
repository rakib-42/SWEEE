import sqlite3
from openpyxl import load_workbook

# Connect to SQLite
conn = sqlite3.connect("database/sweee.db")
cursor = conn.cursor()

# Load workbook
workbook = load_workbook("data/knowledge.xlsx")

# =========================
# Create Tables
# =========================

cursor.execute("""
CREATE TABLE IF NOT EXISTS teachers (
    id TEXT PRIMARY KEY,
    name TEXT,
    designation TEXT,
    department TEXT,
    room TEXT,
    email TEXT,
    office_hours TEXT
)
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS places (
    id TEXT PRIMARY KEY,
    name TEXT,
    building TEXT,
    floor TEXT,
    description TEXT
)
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS robot (
    id TEXT PRIMARY KEY,
    event TEXT,
    speech TEXT,
    motion TEXT,
    expression TEXT
)
""")

# =========================
# Clear old data
# =========================

cursor.execute("DELETE FROM teachers")
cursor.execute("DELETE FROM places")
cursor.execute("DELETE FROM robot")

# =========================
# Import Teachers
# =========================

sheet = workbook["Teachers"]

for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue

    cursor.execute("""
        INSERT INTO teachers
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, row)

# =========================
# Import Places
# =========================

sheet = workbook["Places"]

for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue

    cursor.execute("""
        INSERT INTO places
        VALUES (?, ?, ?, ?, ?)
    """, row)

# =========================
# Import Robot
# =========================

sheet = workbook["Robot"]

for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue

    cursor.execute("""
        INSERT INTO robot
        VALUES (?, ?, ?, ?, ?)
    """, row)

conn.commit()

print("Knowledge imported successfully!")

conn.close()